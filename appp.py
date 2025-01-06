from flask import Flask, render_template, request, send_file, abort, redirect, url_for
import qrcode
from io import BytesIO
from openpyxl import Workbook, load_workbook
import os
import random
from datetime import datetime

app = Flask(__name__)


QUIZ_START_TIME = datetime(2024, 12, 3, 7, 0, 0)  # Example: 7:00 AM
QUIZ_END_TIME = datetime(2024, 12, 3, 22, 0, 0)   # Example: 10:00 PM

AIKEN_FILE = "questions.aiken"


def is_quiz_accessible():
    current_time = datetime.now()
    return QUIZ_START_TIME <= current_time <= QUIZ_END_TIME


def parse_aiken_file(file_path):
    questions = {}
    try:
        with open(file_path, "r") as file:
            lines = file.readlines()

        question_id = 1
        current_question = None
        current_choices = []
        current_answer = None

        for line in lines:
            line = line.strip()
            if line == "":
                continue  # Skip empty lines

            if current_question is None:  # First line of a question
                current_question = line
            elif line[0] in "ABCDEFGHIJKLMNOPQRSTUVWXYZ" and line[1] == ")":  # Choices
                current_choices.append(line[3:])
            elif line.startswith("ANSWER:"):  # Answer line
                current_answer = line.split(":")[1].strip()
                questions[question_id] = {
                    "question": current_question,
                    "choices": current_choices,
                    "answer": [answer.strip() for answer in current_answer.split(',')],
                    "type": "single" if len(current_answer.split(',')) == 1 else "multiple"
                }
                # Reset for next question
                question_id += 1
                current_question = None
                current_choices = []
                current_answer = None

        # Shuffle questions for randomization
        question_ids = list(questions.keys())
        random.shuffle(question_ids)
        return {qid: questions[qid] for qid in question_ids}
    except FileNotFoundError:
        print("Error: Aiken file not found.")
        return {}

# Route: Login Page
@app.route('/')
def login():
    if not is_quiz_accessible():
        return abort(403, "The quiz is not accessible at this time. Check back during the allowed period.")
    return render_template('login.html')

@app.route('/start_quiz', methods=['POST'])
def start_quiz():
    if not is_quiz_accessible():
        return abort(403, "The quiz is not accessible at this time. Check back during the allowed period.")
    name = request.form.get('name')
    email = request.form.get('email')
    roll = request.form.get('roll')
    questions = parse_aiken_file(AIKEN_FILE)

    print("Loaded questions:", questions)  # Debug print to check loaded questions
    return render_template('index.html', name=name, email=email, roll=roll, questions=questions)

def map_answer_to_choice(user_answer, choices):
    """Map the user's answer (like 'Berlin') to the corresponding choice letter ('B')"""
    
    user_answer = user_answer.strip().lower()
    choices = [choice.strip().lower() for choice in choices]  # Normalize all choices
    
    try:
        
        answer_index = choices.index(user_answer)
        print(f"Mapping '{user_answer}' to choice letter: {chr(65 + answer_index)}")  # Debug line
        return chr(65 + answer_index)  # Convert index to corresponding letter (A, B, C, etc.)
    except ValueError:
        # If the answer is not found in choices, return None
        print(f"Answer '{user_answer}' not found in choices")  # Debug line
        return None

@app.route('/submit', methods=['POST'])
def submit():
    name = request.form.get("name")
    email = request.form.get("email")
    roll = request.form.get("roll")

    if not name or not email or not roll:
        return "Error: Missing data", 400

    # Load the questions again when submitting
    questions = parse_aiken_file(AIKEN_FILE)
    print(f"Loaded Questions: {questions}")  # Debug print to check questions

    score = 0  # Initialize score

    for question_id, question_data in questions.items():
        user_answer = request.form.getlist(str(question_id))  # Get answers from the form
        correct_answers = question_data["answer"]  # Correct answers

        print(f"Question ID: {question_id}, User Answer(s): {user_answer}, Correct Answer(s): {correct_answers}")

        # Strip spaces and make answers lowercase to avoid formatting issues
        user_answer = [ans.strip().lower() for ans in user_answer]
        correct_answers = [ans.strip().lower() for ans in correct_answers]

        print(f"User Answers After Stripping: {user_answer}, Correct Answers After Stripping: {correct_answers}")

        if user_answer:  # Check if user_answer is not empty
            if question_data["type"] == "single":
                # Map the user's answer (like 'Paris') to the corresponding choice letter (e.g., 'B')
                user_answer_label = map_answer_to_choice(user_answer[0], question_data['choices'])
                print(f"User Answer Label: {user_answer_label}, Correct Answer(s): {correct_answers}")
                if user_answer_label and user_answer_label.lower() == correct_answers[0].lower():
                    score += 1
            elif question_data["type"] == "multiple":
                
                if sorted(user_answer) == sorted(correct_answers):
                    score += 1

        print(f"Score after question {question_id}: {score}")

   
    save_to_excel(name, email, roll, score)
    print(f"Total score: {score}")  # Debug print to check the final score
    return redirect(url_for('response_saved', name=name))


@app.route('/response_saved')
def response_saved():
    name = request.args.get('name')
    return render_template("response_saved.html", name=name)

def save_to_excel(name, email, roll, score):
    
    print(f"Saving to Excel: Name={name}, Email={email}, Roll={roll}, Score={score}")
    
    filename = "quiz_results.xlsx"
    
    if os.path.exists(filename):
        workbook = load_workbook(filename)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        # Add headers if the file is being created for the first time
        sheet.append(["Name", "Email", "Roll No", "Score"])

    sheet.append([name, email, roll, score])
    workbook.save(filename)


def generate_qr_code(url):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill='black', back_color='white')
    file_path = 'quiz_qr_code.png'
    img.save(file_path)
    return file_path


@app.route('/qr')
def qr_code():
    if not is_quiz_accessible():
        return abort(403, "The quiz is not accessible at this time. Check back during the allowed period.")
    url = 'https://851b-103-160-26-106.ngrok-free.app'  # Replace with your actual quiz URL
    qr_file_path = generate_qr_code(url)
    return f'''
    <img src="/download_qr" alt="QR Code"><p>Scan to access the quiz!</p>
    <a href="/download_qr" download="quiz_qr_code.png">Download QR Code</a>
    '''

# Route: Download QR Code
@app.route('/download_qr')
def download_qr():
    return send_file('quiz_qr_code.png', as_attachment=True)

@app.after_request
def add_header(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001, debug=True)


